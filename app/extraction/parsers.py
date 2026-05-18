from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from app.config import get_settings
from app.extraction.ocr_client import parse_document_with_ocr


IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp", ".tif", ".tiff", ".bmp"}


@dataclass(frozen=True)
class ParsedChunk:
    chunk_index: int
    text: str
    section_name: str | None = None
    page_number: int | None = None
    row_number: int | None = None
    metadata: dict[str, object] = field(default_factory=dict)


@dataclass(frozen=True)
class ParsedDocument:
    raw_text: str
    chunks: list[ParsedChunk]


def _normalize_source_kind(path: Path, source_type: str | None, mime_type: str | None) -> str:
    suffix = path.suffix.lower()
    source_hint = (source_type or "").lower()
    mime_hint = (mime_type or "").lower()

    if source_hint == "pdf" or suffix == ".pdf" or mime_hint == "application/pdf":
        return "pdf"
    if source_hint in {"xlsx", "excel"} or suffix in {".xlsx", ".xlsm"} or "spreadsheetml" in mime_hint:
        return "xlsx"
    if source_hint == "csv" or suffix == ".csv" or mime_hint == "text/csv":
        return "csv"
    if source_hint == "image" or suffix in IMAGE_SUFFIXES or mime_hint.startswith("image/"):
        return "image"
    return "text"


def _single_chunk(raw_text: str, *, section_name: str | None = None) -> ParsedDocument:
    return ParsedDocument(
        raw_text=raw_text,
        chunks=[ParsedChunk(chunk_index=0, text=raw_text, section_name=section_name)],
    )


def _read_text(path: Path) -> ParsedDocument:
    raw_text = path.read_text(encoding="utf-8")
    return _single_chunk(raw_text)


def _row_text(row: dict[str, str]) -> str:
    return " | ".join(f"{key}: {value}" for key, value in row.items() if value not in {None, ""})


def _parse_csv(path: Path) -> ParsedDocument:
    raw_text = path.read_text(encoding="utf-8-sig")
    chunks: list[ParsedChunk] = []
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        for row_index, row in enumerate(reader, start=2):
            normalized_row = {str(key or "column").strip(): str(value or "").strip() for key, value in row.items()}
            text = _row_text(normalized_row)
            if not text:
                continue
            chunks.append(
                ParsedChunk(
                    chunk_index=len(chunks),
                    text=text,
                    section_name=f"Row {row_index}",
                    row_number=row_index,
                    metadata={"parser": "csv"},
                )
            )

    return ParsedDocument(raw_text=raw_text, chunks=chunks or [ParsedChunk(chunk_index=0, text=raw_text, metadata={"parser": "csv"})])


def _format_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_xlsx(path: Path) -> ParsedDocument:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise RuntimeError("openpyxl is required to parse XLSX files") from exc

    workbook = load_workbook(path, data_only=True, read_only=True)
    chunks: list[ParsedChunk] = []
    raw_sections: list[str] = []

    for worksheet in workbook.worksheets:
        header_values: list[str] = []
        header_row_number: int | None = None
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            values = [_format_cell(value) for value in row]
            if not any(values):
                continue
            if not header_values:
                header_values = values
                header_row_number = row_number
                raw_sections.append(f"[{worksheet.title}] " + " | ".join(values))
                continue

            pairs = [
                f"{header_values[index] or f'Column {index + 1}'}: {value}"
                for index, value in enumerate(values)
                if value
            ]
            text = f"{worksheet.title} row {row_number}: " + " | ".join(pairs)
            raw_sections.append(text)
            chunks.append(
                ParsedChunk(
                    chunk_index=len(chunks),
                    text=text,
                    section_name=worksheet.title,
                    row_number=row_number,
                    metadata={
                        "parser": "xlsx",
                        "sheet_name": worksheet.title,
                        "header_row_number": header_row_number or 1,
                    },
                )
            )

    raw_text = "\n".join(raw_sections).strip()
    return ParsedDocument(raw_text=raw_text, chunks=chunks or [ParsedChunk(chunk_index=0, text=raw_text, metadata={"parser": "xlsx"})])


def _decode_pdf_literal(value: str) -> str:
    value = value.replace(r"\(", "(").replace(r"\)", ")").replace(r"\\", "\\")
    value = value.replace(r"\n", "\n").replace(r"\r", "\n").replace(r"\t", "\t")
    return value


def _fallback_pdf_text(path: Path) -> str:
    decoded = path.read_bytes().decode("latin-1", errors="ignore")
    literals = re.findall(r"\(((?:\\.|[^\\)])*)\)\s*Tj", decoded)
    extracted = [_decode_pdf_literal(item.strip()) for item in literals]
    text = "\n".join(value for value in extracted if value).strip()
    if text:
        return text
    printable = "".join(char if char.isprintable() or char in "\n\t" else " " for char in decoded)
    return re.sub(r"\s+", " ", printable).strip()


def _parse_pdf(path: Path) -> ParsedDocument:
    page_texts: list[str] = []
    try:
        from pypdf import PdfReader

        reader = PdfReader(str(path))
        for page in reader.pages:
            page_texts.append((page.extract_text() or "").strip())
    except Exception:  # noqa: BLE001
        page_texts = []

    if not any(page_texts):
        if get_settings().ocr_enabled:
            ocr_document = _parse_ocr(path, source_kind="pdf")
            if ocr_document.raw_text.strip():
                return ocr_document
        fallback_text = _fallback_pdf_text(path)
        page_texts = [fallback_text] if fallback_text else []

    chunks = [
        ParsedChunk(
            chunk_index=index,
            text=text,
            section_name=f"Page {index + 1}",
            page_number=index + 1,
            metadata={"parser": "pdf"},
        )
        for index, text in enumerate(page_texts)
        if text
    ]
    raw_text = "\n\n".join(chunk.text for chunk in chunks)
    return ParsedDocument(raw_text=raw_text, chunks=chunks or [ParsedChunk(chunk_index=0, text=raw_text, metadata={"parser": "pdf"})])


def _parse_ocr(path: Path, *, source_kind: str) -> ParsedDocument:
    settings = get_settings()
    if not settings.ocr_enabled:
        return ParsedDocument(
            raw_text="",
            chunks=[ParsedChunk(chunk_index=0, text="", metadata={"parser": "ocr_disabled", "source_kind": source_kind})],
        )

    try:
        result = parse_document_with_ocr(path)
    except Exception as exc:  # noqa: BLE001 - ingestion should retain the source even if OCR is unavailable
        return ParsedDocument(
            raw_text="",
            chunks=[
                ParsedChunk(
                    chunk_index=0,
                    text="",
                    metadata={"parser": "glm_ocr", "source_kind": source_kind, "ocr_status": "failed", "error": str(exc)[:500]},
                )
            ],
        )

    markdown = str(result.get("markdown") or "").strip()
    page_markdowns = [str(value).strip() for value in list(result.get("page_markdowns") or []) if str(value).strip()]
    chunks = [
        ParsedChunk(
            chunk_index=index,
            text=text,
            section_name=f"OCR page {index + 1}",
            page_number=index + 1,
            metadata={"parser": "glm_ocr", "source_kind": source_kind, "ocr_task_id": str(result.get("task_id") or "")},
        )
        for index, text in enumerate(page_markdowns)
    ]
    if not chunks:
        chunks = [
            ParsedChunk(
                chunk_index=0,
                text=markdown,
                section_name="OCR text",
                metadata={"parser": "glm_ocr", "source_kind": source_kind, "ocr_task_id": str(result.get("task_id") or "")},
            )
        ]
    return ParsedDocument(raw_text=markdown or "\n\n".join(chunk.text for chunk in chunks), chunks=chunks)


def parse_source_file(path: Path, *, source_type: str | None = None, mime_type: str | None = None) -> ParsedDocument:
    source_kind = _normalize_source_kind(path, source_type, mime_type)
    if source_kind == "pdf":
        return _parse_pdf(path)
    if source_kind == "xlsx":
        return _parse_xlsx(path)
    if source_kind == "csv":
        return _parse_csv(path)
    if source_kind == "image":
        return _parse_ocr(path, source_kind="image")
    return _read_text(path)


__all__ = ["ParsedChunk", "ParsedDocument", "parse_source_file"]